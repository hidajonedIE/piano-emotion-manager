
import json
import os
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def generate_translation_xlsx():
    locales_dir = Path(__file__).parent.parent / 'locales'
    
    # Get all languages
    all_languages = sorted([d for d in os.listdir(locales_dir) if os.path.isdir(locales_dir / d)])
    
    # Language names
    language_names = {
        'da': 'Danés',
        'de': 'Alemán',
        'en': 'Inglés',
        'es': 'Español',
        'fr': 'Francés',
        'it': 'Italiano',
        'no': 'Noruego',
        'pt': 'Portugués',
        'sv': 'Sueco'
    }
    
    def flatten_dict(d, parent_key='', sep='.'):
        items = []
        for k, v in d.items():
            new_key = f"{parent_key}{sep}{k}" if parent_key else k
            if isinstance(v, dict):
                items.extend(flatten_dict(v, new_key, sep=sep).items())
            else:
                items.append((new_key, v))
        return dict(items)
    
    # Load all translations
    translations = {}
    for lang in all_languages:
        lang_file = locales_dir / lang / 'translations.json'
        with open(lang_file, 'r', encoding='utf-8') as f:
            translations[lang] = flatten_dict(json.load(f))
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = 'Traducciones'
    
    # Define styles
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Create header
    headers = ['Clave'] + [language_names.get(lang, lang) for lang in all_languages]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Get all keys from Spanish (base language)
    all_keys = sorted(translations['es'].keys())
    
    # Add data rows
    for row_num, key in enumerate(all_keys, 2):
        # Add key
        cell = ws.cell(row=row_num, column=1)
        cell.value = key
        cell.border = border
        cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        
        # Add translations
        for col_num, lang in enumerate(all_languages, 2):
            cell = ws.cell(row=row_num, column=col_num)
            value = translations[lang].get(key, '[FALTA TRADUCCIÓN]')
            cell.value = value
            cell.border = border
            cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 40
    for col_num in range(2, len(all_languages) + 2):
        ws.column_dimensions[chr(64 + col_num)].width = 25
    
    # Freeze header row
    ws.freeze_panes = 'A2'
    
    # Save file
    output_file = locales_dir.parent / 'TRADUCCIONES_CONSOLIDADAS.xlsx'
    wb.save(output_file)
    
    print(f'Archivo XLSX generado: {output_file}')
    print(f'Total de claves: {len(all_keys)}')
    print(f'Total de idiomas: {len(all_languages)}')

if __name__ == '__main__':
    generate_translation_xlsx()
